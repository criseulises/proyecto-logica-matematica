"""
Parser de Archivos (Excel y JSON) - SmartPlannerX
Procesa archivos Excel y JSON con horarios universitarios.

Equipo 3:
- Cristian E. Sánchez R. (25-0688)
- Hansel Augusto Pérez (25-0461)
- Lia De Oleo (25-0673)
- Juan José Cruz Romero (25-0888)
- Samir Gonzalez (25-0808)
- Alejandro Bruno (25-0947)
- Daniel Osvaldo Lopez (25-0655)
"""

import json
import openpyxl
from datetime import datetime
import re


class FileParser:
    """
    Clase para parsear archivos Excel y JSON de horarios universitarios.
    """

    @staticmethod
    def parse_json(file_content):
        """
        Parsea un archivo JSON con estructura de horarios universitarios.

        Returns:
            list: Lista de diccionarios con formato {materia, seccion, dias, inicio, fin}
        """
        try:
            data = json.loads(file_content)
            materias_parsed = []

            # Iterar sobre cada bloque
            for bloque_id, bloque_data in data.items():
                if 'materias' not in bloque_data:
                    continue

                for materia in bloque_data['materias']:
                    codigo = materia.get('codigo', '')
                    nombre = materia.get('nombre', 'Sin Nombre')
                    seccion = materia.get('seccion', '01')
                    horarios = materia.get('horarios', [])

                    # Nombre completo: Código + Nombre
                    nombre_completo = f"{codigo} - {nombre}" if codigo else nombre

                    # Procesar cada horario
                    for horario in horarios:
                        dia = horario.get('dia', '')
                        hora_str = horario.get('hora', '')

                        # Saltar si no hay información válida
                        if not dia or not hora_str or '------' in hora_str:
                            continue

                        # Limpiar hora de marcadores especiales
                        hora_str = hora_str.replace('[ASINCRÓNICA 100%]', '').strip()
                        hora_str = hora_str.replace('VIRTUAL', '').strip()
                        hora_str = hora_str.replace('[HÍBRIDA SINCRÓNICA/ASINCRÓNO 100%]', '').strip()

                        # Extraer horas
                        inicio, fin = FileParser._parse_time_range(hora_str)
                        if inicio is None or fin is None:
                            continue

                        # Mapear día
                        dias = [FileParser._normalize_day(dia)]

                        materias_parsed.append({
                            'materia': nombre_completo,
                            'seccion': str(seccion),
                            'dias': dias,
                            'inicio': inicio,
                            'fin': fin
                        })

            return {'success': True, 'materias': materias_parsed, 'count': len(materias_parsed)}

        except json.JSONDecodeError as e:
            return {'success': False, 'error': f'Error al parsear JSON: {str(e)}'}
        except Exception as e:
            return {'success': False, 'error': f'Error inesperado: {str(e)}'}

    @staticmethod
    def parse_excel(file_path):
        """
        Parsea un archivo Excel con horarios universitarios.

        Returns:
            list: Lista de diccionarios con formato {materia, seccion, dias, inicio, fin}
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            materias_parsed = []
            processed = set()  # Evitar duplicados

            # Intentar leer de la hoja "Resumen General" primero
            sheet_names = ['Resumen General', wb.sheetnames[0]]

            for sheet_name in sheet_names:
                if sheet_name not in wb.sheetnames:
                    continue

                ws = wb[sheet_name]

                # Buscar fila de encabezados
                header_row = None
                for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
                    row_values = [str(cell.value).lower() if cell.value else '' for cell in row]
                    if 'codigo' in row_values or 'materia' in row_values:
                        header_row = idx
                        break

                if not header_row:
                    continue

                # Mapear columnas
                headers = [str(cell.value).lower() if cell.value else '' for cell in ws[header_row]]
                col_map = {}
                for i, header in enumerate(headers):
                    if 'codigo' in header or 'código' in header:
                        col_map['codigo'] = i
                    elif 'materia' in header:
                        col_map['materia'] = i
                    elif 'seccion' in header or 'sección' in header:
                        col_map['seccion'] = i
                    elif 'dia' in header or 'día' in header:
                        col_map['dia'] = i
                    elif 'horario' in header:
                        col_map['horario'] = i

                # Verificar que tengamos las columnas necesarias
                required = ['materia', 'dia', 'horario']
                if not all(col in col_map for col in required):
                    continue

                # Leer datos
                for row in ws.iter_rows(min_row=header_row + 1):
                    try:
                        codigo = str(row[col_map.get('codigo', 0)].value or '').strip()
                        materia = str(row[col_map['materia']].value or '').strip()
                        seccion = str(row[col_map.get('seccion', 0)].value or '01').strip()
                        dia = str(row[col_map['dia']].value or '').strip()
                        horario = str(row[col_map['horario']].value or '').strip()

                        if not materia or not dia or not horario:
                            continue

                        # Nombre completo
                        nombre_completo = f"{codigo} - {materia}" if codigo else materia

                        # Crear clave única para evitar duplicados
                        key = f"{nombre_completo}|{seccion}|{dia}|{horario}"
                        if key in processed:
                            continue
                        processed.add(key)

                        # Parsear horario
                        inicio, fin = FileParser._parse_time_range(horario)
                        if inicio is None or fin is None:
                            continue

                        # Normalizar día
                        dias = [FileParser._normalize_day(dia)]

                        materias_parsed.append({
                            'materia': nombre_completo,
                            'seccion': seccion,
                            'dias': dias,
                            'inicio': inicio,
                            'fin': fin
                        })

                    except Exception as e:
                        continue  # Saltar filas con errores

                # Si encontramos datos, no seguir buscando en otras hojas
                if materias_parsed:
                    break

            wb.close()

            if not materias_parsed:
                return {'success': False, 'error': 'No se encontraron datos válidos en el archivo Excel'}

            return {'success': True, 'materias': materias_parsed, 'count': len(materias_parsed)}

        except Exception as e:
            return {'success': False, 'error': f'Error al leer Excel: {str(e)}'}

    @staticmethod
    def _parse_time_range(time_str):
        """
        Parsea un rango de tiempo del formato "05:00 PM / 08:00 PM" o "5:00 PM / 8:00 PM"

        Returns:
            tuple: (inicio_float, fin_float) o (None, None) si falla
        """
        try:
            # Remover espacios extras y texto adicional
            time_str = time_str.strip()

            # Buscar patrón de hora
            # Formato: HH:MM AM/PM / HH:MM AM/PM
            pattern = r'(\d{1,2}):(\d{2})\s*(AM|PM)\s*/\s*(\d{1,2}):(\d{2})\s*(AM|PM)'
            match = re.search(pattern, time_str, re.IGNORECASE)

            if not match:
                return None, None

            h1, m1, ap1, h2, m2, ap2 = match.groups()

            inicio = FileParser._time_to_float(int(h1), int(m1), ap1.upper())
            fin = FileParser._time_to_float(int(h2), int(m2), ap2.upper())

            return inicio, fin

        except Exception:
            return None, None

    @staticmethod
    def _time_to_float(hours, minutes, ampm):
        """
        Convierte hora en formato 12h a formato decimal 24h.

        Args:
            hours: Horas en formato 12h
            minutes: Minutos
            ampm: 'AM' o 'PM'

        Returns:
            float: Hora en formato decimal 24h
        """
        if ampm == 'PM' and hours != 12:
            hours += 12
        elif ampm == 'AM' and hours == 12:
            hours = 0

        return hours + (minutes / 60.0)

    @staticmethod
    def _normalize_day(day_str):
        """
        Normaliza el nombre del día al formato usado en el sistema.

        Returns:
            str: Día normalizado
        """
        day_map = {
            'LUNES': 'Lunes',
            'MARTES': 'Martes',
            'MIERCOLES': 'Miercoles',
            'MIÉRCOLES': 'Miercoles',
            'JUEVES': 'Jueves',
            'VIERNES': 'Viernes',
            'SABADO': 'Sabado',
            'SÁBADO': 'Sabado',
            'DOMINGO': 'Domingo'
        }

        day_str = day_str.upper().strip()
        return day_map.get(day_str, day_str.capitalize())
